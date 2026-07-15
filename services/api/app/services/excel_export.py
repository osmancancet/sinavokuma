"""Excel çıktıları.

İki ayrı dosya, iki ayrı işi görür:

1) OBS NOT GİRİŞ LİSTESİ (`obs_grade_sheet`)
   OBS'nin "Listeyi Excel'e Aktar" butonuyla indirilen listenin AYNI biçimi.
   Hoca bizden indirir, OBS'ye yükler. Tek tek not girmez.

   Kolonlar OBS ekranından birebir:
       # | Öğrenci No | Girme Durum | Soru1 | Soru2 | ... | Sınav Notu

   OBS kuralları (ekrandaki bilgilendirme metninden):
     - "Girme durumu 'Girmedi' seçilen öğrencilerin not kutucukları kapatılır" → puanlar boş.
     - "İlgili sorunun etki oranından yüksek bir değer girilemez" → puan max_score'u aşamaz.
     - Sınav Notu = soru puanlarının toplamı.

   Ürünün asıl değeri burada: 312 kağıdı okumak değil, 312 satırı OBS'ye elle
   girmemek. Bu dosya olmadan hoca işin yarısını yine elle yapardı.

2) AKREDİTASYON KANIT DOSYASI (`accreditation_workbook`)
   Denetçiye verilecek dosya. DÇ ve PÇ edinim oranları + yöntem açıklaması.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.accreditation import Report

# ── Ortak stiller ────────────────────────────────────────────────────────
LACIVERT = PatternFill("solid", fgColor="16224A")
BASLIK_YAZI = Font(color="FFFFFF", bold=True, size=10)
YESIL = PatternFill("solid", fgColor="D6EDDF")
KIRMIZI = PatternFill("solid", fgColor="FBE0DC")
GRI = PatternFill("solid", fgColor="EFF1F5")
INCE = Side(style="thin", color="D0D6E2")
CERCEVE = Border(left=INCE, right=INCE, top=INCE, bottom=INCE)
ORTA = Alignment(horizontal="center", vertical="center", wrap_text=True)
SAGA = Alignment(horizontal="right")


def _header_row(ws: Worksheet, row: int, headers: list[str], widths: list[int]) -> None:
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = LACIVERT
        cell.font = BASLIK_YAZI
        cell.border = CERCEVE
        cell.alignment = ORTA
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[row].height = 30


# ── 1) OBS NOT GİRİŞ LİSTESİ ─────────────────────────────────────────────


def obs_grade_sheet(
    course_code: str,
    exam_title: str,
    questions: list[dict],  # [{"number": 1, "max_score": 50.0}, ...]
    students: list[dict],  # [{"student_no", "attended", "scores": {q_no: puan}}]
) -> bytes:
    # Ad/soyad kolonu YOK: kağıtta sadece numara var, isim eşleştirmesi OBS'de.
    # OBS bu dosyayı öğrenci numarasından eşleştirip kendi listesine yazar.
    wb = Workbook()
    ws = wb.active
    ws.title = "Not Giriş"

    ws["A1"] = f"{course_code} — {exam_title} · Akreditasyon Not Giriş"
    ws["A1"].font = Font(bold=True, size=13, color="16224A")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + len(questions) + 1)

    ws["A2"] = (
        "Bu dosya OBS'nin not giriş listesiyle aynı biçimdedir. "
        "Puanlar akademisyen tarafından ONAYLANMIŞ notlardır."
    )
    ws["A2"].font = Font(size=9, italic=True, color="5B6684")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3 + len(questions) + 1)

    header_row = 4
    # Sabit kolonlar: # | Öğrenci No | Girme Durum | (sorular...) | Sınav Notu
    FIXED = 3  # numarası: sıra, öğrenci no, girme durum
    headers = ["#", "Öğrenci No", "Girme Durum"]
    widths = [5, 16, 14]
    for q in questions:
        # OBS soru başlıklarında etki oranını gösterir: "Soru1 (%50)"
        headers.append(f"Soru{q['number']} (%{q['max_score']:g})")
        widths.append(13)
    headers.append("Sınav Notu")
    widths.append(12)

    _header_row(ws, header_row, headers, widths)
    note_col = FIXED + len(questions) + 1
    last_col = note_col

    for i, s in enumerate(students, start=1):
        row = header_row + i
        attended = s["attended"]

        ws.cell(row=row, column=1, value=i).alignment = ORTA
        ws.cell(row=row, column=2, value=s["student_no"])

        durum = ws.cell(row=row, column=3, value="Girdi" if attended else "Girmedi")
        durum.alignment = ORTA
        if not attended:
            durum.font = Font(color="8B94AC")

        total = 0.0
        for j, q in enumerate(questions):
            col = FIXED + 1 + j
            cell = ws.cell(row=row, column=col)
            cell.alignment = SAGA
            if not attended:
                # OBS: "Girmedi" seçilen öğrencinin not kutucukları kapatılır.
                cell.fill = GRI
                continue
            score = s["scores"].get(q["number"])
            if score is None:
                continue
            cell.value = round(float(score), 2)
            total += float(score)

        note = ws.cell(row=row, column=note_col, value=0 if not attended else round(total, 2))
        note.alignment = SAGA
        note.font = Font(bold=True)

        for col in range(1, last_col + 1):
            ws.cell(row=row, column=col).border = CERCEVE

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ── 2) AKREDİTASYON KANIT DOSYASI ────────────────────────────────────────


def _attainment_sheet(ws: Worksheet, title: str, items, report: Report) -> None:
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13, color="16224A")
    ws.merge_cells("A1:G1")

    meta = [
        ("Ders", f"{report.course_code} — {report.course_name}"),
        ("Bölüm", report.department or "—"),
        ("Sınav", report.exam_title),
        ("Sınava giren", f"{report.attended_students} / {report.total_students}"),
        ("Hesaba giren (notu onaylanmış)", str(report.approved_students)),
        ("Edinim eşiği", f"%{items[0].threshold:.0f}" if items else "—"),
    ]
    for i, (k, v) in enumerate(meta, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True, size=10)
        ws.cell(row=i, column=2, value=v)

    header_row = 3 + len(meta) + 1
    headers = [
        "Kod",
        "Çıktı Tanımı",
        "İlgili Sorular",
        "Öğrenci",
        "Alınan",
        "Alınabilir",
        "Edinim (%)",
        "Durum",
    ]
    _header_row(ws, header_row, headers, [10, 52, 14, 10, 11, 12, 12, 14])

    for i, a in enumerate(items, start=header_row + 1):
        values = [
            a.code,
            a.description,
            ", ".join(f"S{n}" for n in a.question_numbers),
            a.student_count,
            a.earned,
            a.possible,
            a.pct,
            "EDİNİLDİ" if a.is_attained else "EDİNİLMEDİ",
        ]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.border = CERCEVE
            if col in (4, 5, 6, 7):
                cell.alignment = SAGA
            if col == 2:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 8:
                cell.fill = YESIL if a.is_attained else KIRMIZI
                cell.font = Font(bold=True, size=10)
                cell.alignment = ORTA

    if report.warnings:
        warn_row = header_row + len(items) + 2
        ws.cell(row=warn_row, column=1, value="UYARILAR").font = Font(bold=True, color="BE3A2B")
        for j, w in enumerate(report.warnings, start=1):
            cell = ws.cell(row=warn_row + j, column=1, value=f"• {w}")
            cell.font = Font(color="BE3A2B", size=10)
            ws.merge_cells(
                start_row=warn_row + j, start_column=1, end_row=warn_row + j, end_column=8
            )


YONTEM = """HESAPLAMA YÖNTEMİ

Çıktı zinciri:

    Soru --(ağırlık %)--> DÇ (Ders Öğrenme Çıktısı) --> PÇ (Program Öğrenme Çıktısı)

Bir sorunun bir Ders Öğrenme Çıktısına katkısı:

    katkı = sorunun puanı x o sorudaki DÇ ağırlığı / 100

Örnek: Soru 1 = 50 puan, DÇ1 ağırlığı %25  ->  katkı = 50 x 0,25 = 12,5 puan

Bir DÇ'nin sınıf edinim oranı:

    Edinim (%) = (Alınan / Alınabilir) x 100

    Alınan     : her öğrencinin ilgili sorulardan aldığı puanın oranı x katkı, toplanır
    Alınabilir : katkıların toplamı x değerlendirmeye alınan öğrenci sayısı

Program Öğrenme Çıktısı (PÇ) edinimi, o PÇ'ye bağlı DÇ'lerin katkılarından toplanır.


KANITIN GEÇERLİLİĞİ İÇİN UYGULANAN KURALLAR

1. Yalnızca bir akademisyenin ONAYLADIĞI notlar hesaba katılır.
   Yapay zekânın önerdiği fakat onaylanmamış puanlar bu rapora GİRMEZ.

2. Sınava GİRMEYEN öğrenciler paydaya dahil edilmez.
   Dahil edilseydi sınıfın başarı oranı yapay olarak düşer, rapor yanlış çıkardı.

3. Her puanın yanında, yapay zekânın önerdiği puan ve gerekçesi AYRICA saklanır.
   "Makine ne dedi, insan ne karar verdi" sorusu her kağıt için cevaplanabilir.

4. Hiçbir ders öğrenme çıktısına bağlanmamış sorular hesaba katılmaz ve
   ilgili sayfada UYARI olarak listelenir. Sessizce yutulmaz.

Bu rapor MÜDEK, MEDEK, FEDEK ve YÖKAK kanıt dosyaları için ortak yapıdadır."""


def accreditation_workbook(report: Report) -> bytes:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Program Çıktıları (PÇ)"
    _attainment_sheet(
        ws1, "PROGRAM ÖĞRENME ÇIKTISI EDİNİM RAPORU", report.program_outcomes, report
    )

    ws2 = wb.create_sheet("Ders Çıktıları (DÇ)")
    _attainment_sheet(ws2, "DERS ÖĞRENME ÇIKTISI EDİNİM RAPORU", report.course_outcomes, report)

    ws3 = wb.create_sheet("Yöntem")
    for i, line in enumerate(YONTEM.split("\n"), start=1):
        cell = ws3.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=13, color="16224A")
    ws3.column_dimensions["A"].width = 95

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
